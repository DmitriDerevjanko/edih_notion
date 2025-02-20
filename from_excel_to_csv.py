import os
import csv
from openpyxl import load_workbook

def xlsx_to_csv_with_utf8_bom(xlsx_file_path, csv_file_path=None, sheet_name=None):
    """
    Converts the given XLSX file into a CSV file with UTF-8 BOM (UTF-8-SIG) encoding.

    :param xlsx_file_path: Path to the XLSX file to convert.
    :param csv_file_path:  Path to the output CSV file. If None, it will use the same
                           filename but with a .csv extension.
    :param sheet_name:     Optional sheet name to export. If None, the active sheet is used.
    """
    if not os.path.exists(xlsx_file_path):
        raise FileNotFoundError(f"XLSX file not found: {xlsx_file_path}")

    # If csv_file_path is not specified, construct a default one
    if csv_file_path is None:
        base, _ = os.path.splitext(xlsx_file_path)
        csv_file_path = base + ".csv"

    # Load the Excel workbook
    workbook = load_workbook(xlsx_file_path, read_only=True, data_only=True)

    # Use the specified sheet or the active one
    if sheet_name is None:
        sheet = workbook.active
    else:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in the workbook.")
        sheet = workbook[sheet_name]

    # Create the CSV file with UTF-8 BOM
    with open(csv_file_path, 'w', encoding='utf-8-sig', newline='') as csv_file:
        writer = csv.writer(csv_file, delimiter=',', quoting=csv.QUOTE_MINIMAL)
        
        for row in sheet.iter_rows(values_only=True):
            # row is a tuple of cell values (None if the cell is empty)
            # Convert None values to empty strings (if desired)
            row_data = [(cell if cell is not None else "") for cell in row]
            writer.writerow(row_data)

    print(f"File has been successfully converted to CSV with UTF-8 BOM: {csv_file_path}")

if __name__ == "__main__":
    xlsx_path = "projects_to_update.xlsx"
    try:
        xlsx_to_csv_with_utf8_bom(xlsx_path)
    except Exception as e:
        print(f"Error converting file: {e}")
