import os
import requests
import json
from dotenv import load_dotenv
from datetime import datetime
from dateutil import parser
from openpyxl import Workbook, load_workbook

# Загрузка переменных окружения
load_dotenv()

NOTION_TOKEN = os.getenv("NOTION_TOKEN")
TARGET_DATABASES = json.loads(os.getenv("DATABASES"))

NOTION_API_URL = "https://api.notion.com/v1/databases/{}/query"
HEADERS = {
    "Authorization": f"Bearer {NOTION_TOKEN}",
    "Content-Type": "application/json",
    "Notion-Version": "2022-06-28"
}

def parse_date_to_ymd(date_str: str) -> str:
    """Привести строку к формату YYYY-MM-DD (как текст)."""
    date_str = date_str.strip()
    if not date_str:
        return ""
    try:
        dt = datetime.fromisoformat(date_str)
        return dt.strftime('%Y-%m-%d')
    except ValueError:
        pass
    try:
        dt = parser.parse(date_str)
        return dt.strftime('%Y-%m-%d')
    except (parser.ParserError, ValueError):
        return date_str

def query_notion_database(database_id, start_cursor=None):
    url = NOTION_API_URL.format(database_id)
    data = {}
    if start_cursor:
        data["start_cursor"] = start_cursor

    response = requests.post(url, headers=HEADERS, json=data)
    if response.status_code == 200:
        return response.json()
    else:
        print(f"Failed to query Notion for database {database_id}: {response.status_code}")
        return None

def get_short_description(database_name, customer_name):
    if database_name == 'Digiküpsuse hindamine':
        if 'DMA T1' in customer_name:
            return 'DMA (Digital Maturity Assessment) T1'
        elif 'DMA T0' in customer_name:
            return 'DMA (Digital Maturity Assessment) T0'
    elif database_name == 'AI nõustamine':
        parts = customer_name.split()
        if parts:
            return f'AI suitability analysis {parts[-1]}'
        else:
            return 'AI suitability analysis'
    elif database_name == 'Finantseerimise nõustamine – avalikud meetmed':
        if 'Avalikud meetmed' in customer_name:
            number = customer_name.split()[-1]
            return f'Support to find funding – public measures {number}'
    elif database_name == 'Finantseerimise nõustamine – erakapitali kaasamine':
        if 'Erakapitali kaasamine' in customer_name:
            number = customer_name.split()[-1]
            return f'Consulting to find funding – private investments {number}'
    elif database_name == 'Robotiseerimise nõustamine':
        if 'Robotiseerimise nõustamine' in customer_name:
            number = customer_name.split()[-1]
            return f'Robotics suitability analysis {number}'
    return ''

def get_service_price(database_name):
    """Старое поле Service price."""
    prices = {
        'Digiküpsuse hindamine': '1500',
        'AI nõustamine': '3500',
        'Robotiseerimise nõustamine': '3500',
        'Finantseerimise nõustamine – avalikud meetmed': '2000',
        'Finantseerimise nõustamine – erakapitali kaasamine': '2000'
    }
    return prices.get(database_name, '')

def get_aid_national_price(database_name):
    """
    Логика для нового поля 
    'Amount of the service price to be reported as Aid of national or regional public nature, €':
      - digiküpsuse hindamine, finantseerimise -> 600 (или 800 для finantseerimise, см. ниже)
      - AI nõustamine, robotiseerimise -> 1400
    """
    if database_name == 'Digiküpsuse hindamine':
        return 600
    elif database_name in ('AI nõustamine', 'Robotiseerimise nõustamine'):
        return 1400
    elif database_name.startswith('Finantseerimise nõustamine'):
        # и 'avalikud meetmed', и 'erakapitali kaasamine' → 800
        return 800
    # на всякий случай
    return 0

def find_projects_in_database(database_id, database_name):
    projects = []
    start_cursor = None
    has_more = True

    while has_more:
        data = query_notion_database(database_id, start_cursor)
        if not data:
            break

        for result in data["results"]:
            props = result["properties"]
            service_status = props.get("Service status", {}).get("status", {}).get("name", "")
            edih_status = props.get("EDIH platvormile sisestatud – Finalised", {}).get("select", None)

            # Customer
            project_name = ""
            projekt_field = props.get("Projekt", {}).get("title", [])
            if projekt_field:
                project_name = projekt_field[0].get("text", {}).get("content", "")

            # VAT
            vat_value = ""
            # Если digiküpsuse hindamine - берём из "Registrikood" (числовое поле)
            if database_name == 'Digiküpsuse hindamine':
                vat_num = props.get("Registrikood", {}).get("number", None)
                if vat_num is not None:
                    vat_value = str(vat_num)
            else:
                # Предыдущая логика c rollup:
                property_data = props.get("Registrikood, automaatne lahter, lohista alla")
                if not property_data:
                    # fallback
                    property_data = props.get("Registrikood", {})
                
                rollup_info = property_data.get("rollup", {})
                if rollup_info.get("type") == "array":
                    arr = rollup_info.get("array", [])
                    if arr:
                        first_item = arr[0]
                        if first_item.get("type") == "number":
                            n_val = first_item.get("number")
                            if n_val is not None:
                                vat_value = str(n_val)

            # Dates
            if database_name == 'Digiküpsuse hindamine':
                start_field = props.get("Automaatne väli, DMA link ettevõttele saadetud (teenuse algus)")
                start_date_raw = ""
                if start_field and start_field.get("date"):
                    start_date_raw = start_field["date"].get("start", "")
            else:
                start_field = props.get("Esmanõustamise kuupäev (ev külastuse kpv, teenuse osutamise alguse kpv)")
                start_date_raw = ""
                if start_field and start_field.get("rich_text"):
                    start_date_raw = start_field["rich_text"][0].get("text", {}).get("content", "")

            finish_field = props.get("Raport valminud - nõustamine tehtud, võib VTA välja maksta")
            finish_date_raw = ""
            if finish_field and finish_field.get("date"):
                finish_date_raw = finish_field["date"].get("start", "")

            # Short description
            short_description = get_short_description(database_name, project_name)

            # Условие: service_status == "Finalised" и edih_status is None
            if service_status == "Finalised" and edih_status is None:
                projects.append((
                    project_name,
                    vat_value,
                    short_description,
                    start_date_raw,
                    finish_date_raw
                ))

        start_cursor = data.get("next_cursor", None)
        has_more = data.get("has_more", False)

    return projects

# Новая функция для загрузки маппинга VAT->CompanyName из export-sme.xlsx
def load_sme_mapping(filepath="export-sme.xlsx"):
    """
    Считывает файл `export-sme.xlsx`.
    Предполагаем, что:
      - столбец D (4-й) содержит VAT (registrikood);
      - столбец B (2-й) содержит название компании.
    Возвращает словарь вида { 'someVAT': 'CompanyName', ... }
    """
    vat_to_company = {}
    wb = load_workbook(filepath)
    ws = wb.active  # Если нужный лист не первый, укажите по имени или индексу
    for row in ws.iter_rows(min_row=2, values_only=True):
        # Предположим, что row = (A, B, C, D, ...)
        company_name = row[1]  # столбец B
        vat_number = row[3]    # столбец D
        if vat_number is not None:
            # Сохраняем в словарь
            vat_to_company[str(vat_number).strip()] = company_name if company_name else ""
    return vat_to_company

def save_to_excel(all_projects):
    # Загрузим справочник VAT->Название из export-sme.xlsx
    vat_map = load_sme_mapping("export-sme.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "Projects to Update"

    # Добавляем новый заголовок (Aid of national...) после "Price invoiced to customer"
    headers = [
        "Content ID",
        "Customer",
        "VAT",
        "Service category delivered",
        "Number of attendees",
        "Service price",
        "Price invoiced to customer",
        "Amount of the service price to be reported as Aid of national or regional public nature, €",
        "Specific information on State Aid",
        "Technology type used",
        "Status",
        "Short description of the service",
        "Amount of investment triggered",
        "Type of investment",
        "Dates",
        "Information on the use of capacities financed by the Digital Europe Programme",
        "Specify use of capacities financed by the Digital Europe Programme",
        "Specify type of investment",
        "Specify Technology type used",
        "Customer ID",
        "EDIH Name",
        "EDIH Country",
        "Customer Country",
        "Customer  region",
        "Customer primary sector",
        "Customer staff size",
        "Customer type",
        "Author's email"
    ]
    ws.append(headers)

    for database_name, projects in all_projects.items():
        for project_name, vat_value, short_description, start_date_raw, finish_date_raw in projects:
            # Service price
            sp_str = get_service_price(database_name)
            try:
                service_price_num = int(sp_str) if sp_str else 0
            except ValueError:
                service_price_num = 0

            # Price invoiced to customer
            price_invoiced_num = 0

            # Amount of the service price for Aid
            aid_price_num = get_aid_national_price(database_name)

            # Service category
            if database_name in (
                'Finantseerimise nõustamine – avalikud meetmed',
                'Finantseerimise nõustamine – erakapitali kaasamine'
            ):
                service_category = 'Support to find investment'
            else:
                service_category = 'Test before invest'

            # Даты в формат YYYY-MM-DD
            start_date_ymd = parse_date_to_ymd(start_date_raw)
            finish_date_ymd = parse_date_to_ymd(finish_date_raw)
            if start_date_ymd or finish_date_ymd:
                dates_text = f"{start_date_ymd} / {finish_date_ymd}".strip()
            else:
                dates_text = ""

            # Проверяем в словаре VAT->Название
            # Если есть, берём название, если нет, "error"
            matched_company_name = vat_map.get(vat_value.strip(), "error")

            row_data = [
                "",
                matched_company_name,              # Customer (либо название, либо "error")
                vat_value,                         # VAT
                service_category,                  # Service category
                "",
                service_price_num,
                price_invoiced_num,
                aid_price_num,
                "",
                "Artificial Intelligence & Decision support;Robotics",
                'Finalised and invoiced',
                short_description,
                "",
                "",
                dates_text,
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                ""
            ]
            ws.append(row_data)

    file_path = "projects_to_update.xlsx"
    wb.save(file_path)
    print(f"Excel file saved as: {file_path}")

def check_all_databases():
    all_projects_to_update = {}
    for database_id, database_name in TARGET_DATABASES.items():
        print(f"Checking database: {database_name} (ID: {database_id})")
        projects_to_update = find_projects_in_database(database_id, database_name)
        if projects_to_update:
            all_projects_to_update[database_name] = projects_to_update
            print(f"Projects to update in {database_name}:")
            for proj, vat, desc, start_date, end_date in projects_to_update:
                print(f"- {proj} (VAT: {vat}, {desc}) - Start: {start_date}, Finish: {end_date}")
        else:
            print(f"No projects found that require updates in {database_name}.")

    return all_projects_to_update

if __name__ == "__main__":
    all_projects = check_all_databases()
    if all_projects:
        save_to_excel(all_projects)
    else:
        print("No projects found that require updates in any database.")
